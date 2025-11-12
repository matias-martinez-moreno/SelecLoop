# =============================================================================
# VISTAS DE LA APLICACIÓN COMPANIES - SelecLoop
# =============================================================================
# Este archivo contiene las vistas relacionadas con empresas
#
# Vistas principales:
# - company_detail_view: Vista detallada de una empresa
# - company_dashboard_view: Dashboard para representantes de empresa
# =============================================================================

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg, Count
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from .models import Company
from accounts.models import UserProfile
from .forms import CompanyEditForm


@login_required
def company_detail_view(request, company_id):
    """Vista detallada de una empresa"""
    # Staff puede ver empresas inactivas, otros usuarios solo activas
    if request.user.is_staff or (hasattr(request.user, 'profile') and request.user.profile.role == 'staff'):
        company = get_object_or_404(Company, id=company_id)
    else:
        company = get_object_or_404(Company, id=company_id, is_active=True)
    
    # Importar modelos necesarios
    from reviews.models import Review, PendingReview
    
    # Verificar acceso del usuario
    user_can_access = False
    user_has_contributed = False
    user_review_status = None
    user_has_pending_review = False
    user_can_create_review = False
    
    # Staff puede ver todo sin restricciones
    if request.user.is_staff or (hasattr(request.user, 'profile') and request.user.profile.role == 'staff'):
        user_can_access = True
        user_can_create_review = False  # Los staff no pueden crear reseñas
    elif request.user.profile.role == 'company_rep':
        user_can_access = True
        user_can_create_review = False  # Los company_rep no pueden crear reseñas
    elif request.user.profile.role == 'candidate':
        # Verificar si tiene reseñas pendientes
        any_pending_reviews = PendingReview.objects.filter(
            user_profile=request.user.profile,
            is_reviewed=False
        ).exists()
        
        if any_pending_reviews:
            user_can_access = False
            user_has_pending_review = True
        else:
            user_can_access = True
            user_can_create_review = True  # Los candidatos pueden crear reseñas
            
            # Verificar reseñas para esta empresa
            user_reviews = Review.objects.filter(
                user_profile=request.user.profile,
                company=company
            )
            
            if user_reviews.exists():
                user_has_contributed = True
                latest_review = user_reviews.latest('submission_date')
                user_review_status = latest_review.status
    
    # Obtener todas las reseñas aprobadas SIN filtrar (para contar el total)
    all_approved_reviews = Review.objects.filter(
        company=company,
        status='approved'
    ).select_related('user_profile__user')
    
    # Obtener reseñas con filtros (se aplicarán después)
    approved_reviews = all_approved_reviews
    
    # Aplicar filtros desde parámetros GET
    rating_filter = request.GET.get('rating')
    modality_filter = request.GET.get('modality')
    sort_by = request.GET.get('sort', 'recent')  # Por defecto: más recientes
    
    if rating_filter:
        try:
            rating_value = int(rating_filter)
            if 1 <= rating_value <= 5:
                approved_reviews = approved_reviews.filter(overall_rating=rating_value)
        except ValueError:
            pass
    
    if modality_filter:
        approved_reviews = approved_reviews.filter(modality=modality_filter)
    
    # Aplicar ordenamiento
    if sort_by == 'recent':
        approved_reviews = approved_reviews.order_by('-submission_date')
    elif sort_by == 'oldest':
        approved_reviews = approved_reviews.order_by('submission_date')
    elif sort_by == 'highest':
        approved_reviews = approved_reviews.order_by('-overall_rating', '-submission_date')
    elif sort_by == 'lowest':
        approved_reviews = approved_reviews.order_by('overall_rating', '-submission_date')
    else:
        approved_reviews = approved_reviews.order_by('-submission_date')
    
    rejected_reviews = Review.objects.filter(
        company=company,
        status='rejected'
    ).select_related('user_profile__user')
    
    # Combinar según rol - solo mostrar aprobadas para candidatos, todas para empresas y staff
    if request.user.profile.role == 'company_rep' or request.user.is_staff or (hasattr(request.user, 'profile') and request.user.profile.role == 'staff'):
        all_visible_reviews = list(approved_reviews) + list(rejected_reviews)
    else:
        all_visible_reviews = list(approved_reviews)  # Solo aprobadas para candidatos
    
    # Reseña pendiente específica
    pending_review = None
    if request.user.profile.role == 'candidate':
        pending_review = PendingReview.objects.filter(
            user_profile=request.user.profile,
            company=company,
            is_reviewed=False
        ).first()
    
    # ===== Estadísticas y datos para gráficos =====
    # Para reputación base, por defecto usamos reseñas aprobadas.
    if hasattr(request.user, 'profile') and request.user.profile.role == 'candidate':
        reviews_for_stats = Review.objects.filter(company=company, status='approved')
    elif request.user.is_staff or (hasattr(request.user, 'profile') and request.user.profile.role == 'company_rep'):
        reviews_for_stats = Review.objects.filter(company=company)
    else:
        reviews_for_stats = approved_reviews

    # Promedios numéricos
    from django.db.models import Avg as DJAvg, Count as DJCount
    from django.utils import timezone
    from datetime import timedelta

    avg_overall = reviews_for_stats.aggregate(v=DJAvg('overall_rating'))['v'] or 0

    # Mapear categorías a valores para promediar
    COMM_SCORES = {
        'excellent': 5,
        'good': 4,
        'regular': 3,
        'poor': 2,
    }
    DIFF_SCORES = {
        'very_easy': 1,
        'easy': 2,
        'moderate': 3,
        'difficult': 4,
        'very_difficult': 5,
    }
    RESP_SCORES = {
        'immediate': 5,
        'same_day': 4,
        'next_day': 3,
        'few_days': 2,
        'slow': 1,
    }

    def avg_from_choices(qs, field, scores):
        total = 0
        count = 0
        for val, c in qs.values_list(field).annotate(cnt=DJCount('pk')):
            if val in scores:
                total += scores[val] * qs.filter(**{field: val}).count()
                count += qs.filter(**{field: val}).count()
        return (total / count) if count else 0

    avg_communication = avg_from_choices(reviews_for_stats, 'communication_rating', COMM_SCORES)
    avg_difficulty = avg_from_choices(reviews_for_stats, 'difficulty_rating', DIFF_SCORES)
    avg_response_time = avg_from_choices(reviews_for_stats, 'response_time_rating', RESP_SCORES)

    total_reviews_stats = reviews_for_stats.count()
    company_stats = {
        'avg_overall': avg_overall,
        'avg_communication': avg_communication,
        'avg_difficulty': avg_difficulty,
        'avg_response_time': avg_response_time,
        'total_reviews': total_reviews_stats,
    }

    # ===== DATOS PARA GRÁFICOS =====
    chart_data = {
        'ratings': {},
        'modality': {},
        'status': {},
        'timeline': {},
        'strengths_weaknesses': {}
    }
    
    if reviews_for_stats.exists():
        # Datos de distribución de calificaciones
        rating_counts = {}
        for i in range(1, 6):
            count = reviews_for_stats.filter(overall_rating=i).count()
            rating_counts[f'{i} estrella{"s" if i > 1 else ""}'] = count
        chart_data['ratings'] = rating_counts
        
        # Datos de modalidad
        modality_counts = {}
        for modality, _ in Review.MODALITY_CHOICES:
            count = reviews_for_stats.filter(modality=modality).count()
            if count > 0:
                modality_counts[modality.title()] = count
        chart_data['modality'] = modality_counts
        
        # Datos de estado
        status_counts = {}
        for status, _ in Review.STATUS_CHOICES:
            count = reviews_for_stats.filter(status=status).count()
            if count > 0:
                status_counts[status.title()] = count
        chart_data['status'] = status_counts
        
        # Datos de timeline - obtener todos los meses donde hay reseñas
        timeline_data = {}
        from django.db.models.functions import TruncMonth
        
        # Obtener todos los meses únicos donde hay reseñas
        timeline_months = reviews_for_stats.annotate(
            month=TruncMonth('submission_date')
        ).values('month').annotate(
            count=DJCount('id')
        ).order_by('month')
        
        # Crear diccionario con formato de fecha legible
        for item in timeline_months:
            month_date = item['month']
            if month_date:
                month_name = month_date.strftime('%b %Y')
                timeline_data[month_name] = item['count']
        
        chart_data['timeline'] = timeline_data
        
        # Debug: Agregar información adicional para verificar
        chart_data['debug_info'] = {
            'total_reviews': reviews_for_stats.count(),
            'latest_review_date': reviews_for_stats.order_by('-submission_date').first().submission_date if reviews_for_stats.exists() else None,
            'current_time': timezone.now().isoformat(),
            'ratings_data': chart_data['ratings'],
            'modality_data': chart_data['modality'],
            'status_data': chart_data['status'],
            'timeline_data': chart_data['timeline'],
            'strengths_data': chart_data['strengths_weaknesses']
        }
        
        # Gráfico de Fortalezas y Debilidades (Radar Chart)
        strengths_weaknesses = {}
        from django.db.models import Avg
        
        # Calcular promedios por aspecto (usando solo campos disponibles)
        aspects = {
            'communication_rating': 'Comunicación',
            'difficulty_rating': 'Dificultad del Proceso', 
            'response_time_rating': 'Tiempo de Respuesta',
            'overall_rating': 'Calificación General'
        }
        
        for field, label in aspects.items():
            if field == 'overall_rating':
                # Para overall_rating usamos Avg directamente
                avg_rating = reviews_for_stats.aggregate(
                    avg=Avg(field)
                )['avg']
                if avg_rating is not None:
                    strengths_weaknesses[label] = round(avg_rating, 1)
            else:
                # Para los otros campos, necesitamos convertir las opciones a números
                # Mapeo de opciones a valores numéricos
                rating_mapping = {
                    'communication_rating': {
                        'excellent': 5, 'good': 4, 'regular': 3, 'poor': 2
                    },
                    'difficulty_rating': {
                        'very_easy': 1, 'easy': 2, 'moderate': 3, 'difficult': 4, 'very_difficult': 5
                    },
                    'response_time_rating': {
                        'immediate': 5, 'same_day': 4, 'next_day': 3, 'few_days': 2, 'slow': 1
                    }
                }
                
                # Calcular promedio manualmente
                total_score = 0
                count = 0
                for review in reviews_for_stats:
                    rating_value = rating_mapping[field].get(getattr(review, field), 0)
                    if rating_value > 0:
                        total_score += rating_value
                        count += 1
                
                if count > 0:
                    avg_rating = total_score / count
                    strengths_weaknesses[label] = round(avg_rating, 1)
        
        chart_data['strengths_weaknesses'] = strengths_weaknesses
        
        # Datos de distribución de tiempo de respuesta
        response_time_counts = {}
        response_time_labels = {
            'immediate': 'Inmediata',
            'same_day': 'Mismo día',
            'next_day': 'Al día siguiente',
            'few_days': 'En pocos días',
            'slow': 'Lenta'
        }
        for response_time, _ in Review.RESPONSE_TIME_CHOICES:
            count = reviews_for_stats.filter(response_time_rating=response_time).count()
            if count > 0:
                response_time_counts[response_time_labels[response_time]] = count
        chart_data['response_time_distribution'] = response_time_counts
    
    # ===== Estadísticas específicas por rol =====
    role_kpis = {}
    try:
        user_role = request.user.profile.role
    except Exception:
        user_role = None

    if user_role == 'candidate':
        # Estadísticas útiles para candidatos
        since_90 = timezone.now() - timedelta(days=90)
        last90_reviews = reviews_for_stats.filter(submission_date__gte=since_90)

        # Tasa de respuesta rápida (últimos 90 días)
        fast_responses = last90_reviews.filter(response_time_rating__in=['immediate', 'same_day']).count()
        fast_response_rate = (fast_responses / last90_reviews.count()) * 100 if last90_reviews.count() > 0 else 0

        # Modalidad más recomendada
        top_modality = reviews_for_stats.values('modality').annotate(
            count=DJCount('id'),
            avg_rating=Avg('overall_rating')
        ).order_by('-avg_rating', '-count').first()

        # Calcular promedios para todas las modalidades (para rotación en el mensaje)
        # Asegurar que siempre incluimos las 3 modalidades principales si tienen datos
        all_modalities = reviews_for_stats.values('modality').annotate(
            count=DJCount('id'),
            avg_rating=Avg('overall_rating')
        ).order_by('modality')
        
        modalities_data = []
        # Mapeo de modalidades para normalizar nombres (remoto = online)
        modality_map = {
            'presencial': 'presencial',
            'remoto': 'online',
            'online': 'online',
            'hibrido': 'híbrido',
            'hybrid': 'híbrido',
            'híbrido': 'híbrido'
        }
        
        for mod in all_modalities:
            if mod['modality'] and mod['count'] > 0:
                # Normalizar el nombre de la modalidad
                normalized_modality = modality_map.get(mod['modality'].lower(), mod['modality'].lower())
                modalities_data.append({
                    'modality': normalized_modality,
                    'avg_rating': round(mod['avg_rating'], 1),
                    'count': mod['count']
                })
        
        # Ordenar por orden específico: presencial, online, híbrido
        modality_order = {'presencial': 0, 'online': 1, 'híbrido': 2}
        modalities_data.sort(key=lambda x: modality_order.get(x['modality'], 999))

        # Nivel de dificultad promedio
        difficulty_distribution = reviews_for_stats.values('difficulty_rating').annotate(
            count=DJCount('id')
        ).order_by('difficulty_rating')

        # Calidad de comunicación
        comm_distribution = reviews_for_stats.values('communication_rating').annotate(
            count=DJCount('id')
        ).order_by('communication_rating')
        
        # Calcular promedios para las métricas de calidad (usar la misma función que más arriba)
        avg_communication = avg_from_choices(reviews_for_stats, 'communication_rating', COMM_SCORES)
        avg_difficulty = avg_from_choices(reviews_for_stats, 'difficulty_rating', DIFF_SCORES)
        avg_response_time = avg_from_choices(reviews_for_stats, 'response_time_rating', RESP_SCORES)

        role_kpis = {
            'role': 'candidate',
            'avg_overall': avg_overall,
            'last90_reviews_count': last90_reviews.count(),
            'fast_response_rate': round(fast_response_rate, 1),
            'top_modality': top_modality['modality'] if top_modality else None,
            'top_modality_rating': round(top_modality['avg_rating'], 1) if top_modality else 0,
            'modalities_data': modalities_data,  # Datos de todas las modalidades para rotación
            'difficulty_distribution': list(difficulty_distribution),
            'communication_distribution': list(comm_distribution),
            'total_reviews': reviews_for_stats.count(),
            'avg_communication': avg_communication,  # Promedio de comunicación (1-5)
            'avg_difficulty': avg_difficulty,  # Promedio de dificultad (1-5)
            'avg_response_time': avg_response_time,  # Promedio de tiempo de respuesta (1-5)
        }
    elif user_role == 'company_rep':
        # Estadísticas útiles para representantes de empresa
        all_company_reviews = Review.objects.filter(company=company)

        # Ratios de aprobación
        total_reviews = all_company_reviews.count()
        approved_count = all_company_reviews.filter(status='approved').count()
        rejected_count = all_company_reviews.filter(status='rejected').count()

        approval_rate = (approved_count / total_reviews) * 100 if total_reviews > 0 else 0
        rejection_rate = (rejected_count / total_reviews) * 100 if total_reviews > 0 else 0

        # Compromiso de tiempo de respuesta (aprobadas con respuesta rápida)
        compromiso_compliant = all_company_reviews.filter(
            status='approved',
            response_time_rating__in=['immediate', 'same_day']
        ).count()
        compromiso_rate = (compromiso_compliant / approved_count) * 100 if approved_count > 0 else 0

        # Tendencia mensual - obtener todos los meses donde hay reseñas
        # Obtener todos los meses únicos donde hay reseñas, sin límite de tiempo
        monthly_trend = all_company_reviews.annotate(
            month=TruncMonth('submission_date')
        ).values('month').annotate(
            count=DJCount('id'),
            approved=DJCount('id', filter=Q(status='approved')),
            avg_rating=DJAvg('overall_rating')
        ).order_by('month')
        
        # Calificaciones por mes
        monthly_ratings = list(monthly_trend)
        
        # Agregar datos mensuales al chart_data para el gráfico
        if monthly_ratings:
            chart_data['monthly_comparison'] = {
                str(item['month']): {
                    'count': item['count'],
                    'approved': item['approved'],
                    'avg_rating': float(item['avg_rating']) if item['avg_rating'] else 0
                }
                for item in monthly_ratings
            }

        # Generar recomendaciones inteligentes
        recommendations = []
        
        if avg_communication < 4:
            recommendations.append({
                'type': 'warning',
                'icon': 'fas fa-comments',
                'title': 'Mejorar Comunicación',
                'description': f'Tu calificación de comunicación es {avg_communication:.1f}/5. Considera ser más claro y transparente en tus respuestas a los candidatos.',
                'action': 'Revisa tus procesos de comunicación y asegúrate de dar feedback claro y oportuno.'
            })
        
        if avg_response_time < 4:
            recommendations.append({
                'type': 'info',
                'icon': 'fas fa-clock',
                'title': 'Acelerar Respuestas',
                'description': f'Tu tiempo de respuesta promedio es {avg_response_time:.1f}/5. Los candidatos valoran respuestas rápidas.',
                'action': 'Establece un SLA de respuesta de máximo 24 horas para dar mejor impresión.'
            })
        
        if rejection_rate > 30:
            recommendations.append({
                'type': 'danger',
                'icon': 'fas fa-exclamation-triangle',
                'title': 'Alta Tasa de Rechazo',
                'description': f'El {rejection_rate:.1f}% de tus reseñas son rechazadas. Esto puede indicar problemas en tu proceso.',
                'action': 'Revisa las reseñas rechazadas para entender qué está causando el rechazo automático.'
            })
        
        if approval_rate < 70:
            recommendations.append({
                'type': 'warning',
                'icon': 'fas fa-chart-line',
                'title': 'Baja Tasa de Aprobación',
                'description': f'Solo el {approval_rate:.1f}% de tus reseñas son aprobadas. Mejora la calidad de tus procesos.',
                'action': 'Analiza las reseñas aprobadas para replicar lo que funciona bien.'
            })
        
        if compromiso_rate < 50:
            recommendations.append({
                'type': 'info',
                'icon': 'fas fa-bolt',
                'title': 'Mejorar Velocidad de Respuesta',
                'description': f'Solo el {compromiso_rate:.1f}% de tus reseñas aprobadas tienen respuesta rápida.',
                'action': 'Implementa respuestas automáticas y establece procesos más ágiles.'
            })
        
        if avg_overall < 3.5:
            recommendations.append({
                'type': 'danger',
                'icon': 'fas fa-star',
                'title': 'Calificación General Baja',
                'description': f'Tu calificación promedio es {avg_overall:.1f}/5. Necesitas mejorar varios aspectos.',
                'action': 'Revisa todas las áreas de mejora y crea un plan de acción prioritario.'
            })
        
        if not recommendations:
            recommendations.append({
                'type': 'success',
                'icon': 'fas fa-check-circle',
                'title': '¡Excelente Trabajo!',
                'description': 'Tus métricas están en buen nivel. Sigue así y continúa mejorando.',
                'action': 'Mantén la calidad y considera implementar mejoras incrementales.'
            })
        
        # Datos para comparación mes a mes
        monthly_comparison = []
        if len(monthly_ratings) >= 2:
            for i in range(1, len(monthly_ratings)):
                prev_month = monthly_ratings[i-1]
                curr_month = monthly_ratings[i]
                
                # Calcular rechazadas para cada mes
                prev_rejected = prev_month['count'] - prev_month['approved']
                curr_rejected = curr_month['count'] - curr_month['approved']
                
                rating_change = curr_month['avg_rating'] - prev_month['avg_rating']
                count_change = curr_month['count'] - prev_month['count']
                approval_change = curr_month['approved'] - prev_month['approved']
                rejection_change = curr_rejected - prev_rejected
                
                monthly_comparison.append({
                    'month': curr_month['month'],
                    'prev_month': prev_month['month'],  # Mes anterior para mostrar
                    'prev_rating': prev_month['avg_rating'],
                    'curr_rating': curr_month['avg_rating'],
                    'rating_change': rating_change,
                    'prev_count': prev_month['count'],
                    'curr_count': curr_month['count'],
                    'count_change': count_change,
                    'prev_approved': prev_month['approved'],
                    'curr_approved': curr_month['approved'],
                    'approval_change': approval_change,
                    'prev_rejected': prev_rejected,
                    'curr_rejected': curr_rejected,
                    'rejection_change': rejection_change,
                })
        
        role_kpis = {
            'role': 'company_rep',
            'avg_overall': avg_overall,
            'total_reviews': total_reviews,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'approval_rate': round(approval_rate, 1),
            'rejection_rate': round(rejection_rate, 1),
            'compromiso_rate': round(compromiso_rate, 1),
            'monthly_trend': monthly_ratings,
            'monthly_comparison': monthly_comparison,
            'avg_communication': avg_communication,
            'avg_difficulty': avg_difficulty,
            'avg_response_time': avg_response_time,
            'recommendations': recommendations,
    }

    # Contar el total de reseñas aprobadas SIN filtrar
    total_approved_reviews_count = all_approved_reviews.count()
    
    # Verificar si hay filtros activos (excluyendo el ordenamiento por defecto 'recent')
    has_active_filters = bool(rating_filter or modality_filter)
    
    # Pasar filtros actuales al contexto
    current_filters = {
        'rating': rating_filter,
        'modality': modality_filter,
        'sort': sort_by,
    }
    
    context = {
        'company': company,
        'approved_reviews': approved_reviews,
        'total_approved_reviews_count': total_approved_reviews_count,  # Total sin filtrar
        'has_active_filters': has_active_filters,  # Si hay filtros activos
        'rejected_reviews': rejected_reviews,
        'all_visible_reviews': all_visible_reviews,
        'user_can_access': user_can_access,
        'user_has_contributed': user_has_contributed,
        'user_review_status': user_review_status,
        'user_has_pending_review': user_has_pending_review,
        'user_can_create_review': user_can_create_review,
        'pending_review': pending_review,
        'company_stats': company_stats,
        'chart_data': chart_data,
        'role_kpis': role_kpis,
        'current_filters': current_filters,
    }
    
    return render(request, 'core/company_detail.html', context)


@login_required
def export_company_report_excel(request, company_id):
    """Exportar reporte de empresa a Excel (.xlsx) con formato profesional"""
    company = get_object_or_404(Company, id=company_id)
    
    # Verificar permisos
    if not (request.user.is_staff or request.user.profile.role == 'company_rep'):
        messages.error(request, '❌ No tienes permisos para exportar reportes.')
        return redirect('dashboard')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.chart.series import Series
    except ImportError:
        messages.error(request, '⚠️ La librería openpyxl no está instalada. Instálala con: pip install openpyxl')
        return redirect('company_detail', company_id=company_id)
    
    from reviews.models import Review
    from django.db.models import Avg, Count, Q
    from datetime import datetime
    from io import BytesIO
    import locale
    
    # Configurar locale en español para las fechas
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, 'es_CO.UTF-8')
        except locale.Error:
            # Si no hay locale español disponible, usar formato manual
            pass
    
    # Obtener período seleccionado
    period = request.GET.get('period', 'all')
    period_label = 'Todas las reseñas'
    
    # Mapeo de meses en español
    meses_espanol = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    
    if period != 'all':
        # Verificar si es un trimestre (formato: Q1-2025, Q2-2025, etc.)
        if period.startswith('Q') and '-' in period:
            try:
                quarter_str, year_str = period.split('-')
                quarter = int(quarter_str[1:])  # Extraer el número del trimestre (1-4)
                year_int = int(year_str)
                
                # Determinar los meses del trimestre
                if quarter == 1:
                    months = [1, 2, 3]
                    period_label = f'Q1 {year_int} (Ene-Mar)'
                elif quarter == 2:
                    months = [4, 5, 6]
                    period_label = f'Q2 {year_int} (Abr-Jun)'
                elif quarter == 3:
                    months = [7, 8, 9]
                    period_label = f'Q3 {year_int} (Jul-Sep)'
                elif quarter == 4:
                    months = [10, 11, 12]
                    period_label = f'Q4 {year_int} (Oct-Dic)'
                else:
                    raise ValueError("Trimestre inválido")
                
                all_reviews = Review.objects.filter(
                    company=company,
                    submission_date__year=year_int,
                    submission_date__month__in=months
                ).select_related('user_profile__user')
            except (ValueError, AttributeError):
                period = 'all'
                all_reviews = Review.objects.filter(company=company).select_related('user_profile__user')
        else:
            # Es un mes individual (formato: 2025-07)
            try:
                year, month = period.split('-')
                year_int = int(year)
                month_int = int(month)
                # Formatear en español
                month_name = meses_espanol.get(month_int, '')
                period_label = f'{month_name.capitalize()} {year_int}'
                all_reviews = Review.objects.filter(
                    company=company,
                    submission_date__year=year,
                    submission_date__month=month
                ).select_related('user_profile__user')
            except (ValueError, AttributeError):
                period = 'all'
                all_reviews = Review.objects.filter(company=company).select_related('user_profile__user')
    else:
        all_reviews = Review.objects.filter(company=company).select_related('user_profile__user')
    
    # Filtrar por calificación si se especifica (solo para reseñas aprobadas)
    rating_filter = request.GET.get('rating')
    rating_label = ''
    if rating_filter:
        try:
            rating_value = int(rating_filter)
            if 1 <= rating_value <= 5:
                rating_label = f' - Calificación {rating_value} estrellas'
        except ValueError:
            pass
    
    # Obtener columnas seleccionadas
    selected_columns = request.GET.getlist('columns')
    if not selected_columns:
        # Si no se seleccionaron columnas, incluir todas por defecto
        selected_columns = ['fecha', 'usuario', 'cargo', 'modalidad', 'calificacion', 
                           'comunicacion', 'dificultad', 'tiempo_respuesta', 'pros', 'contras', 'preguntas']
    
    # Mapeo de columnas
    column_map = {
        'fecha': ('Fecha', 'A'),
        'usuario': ('Usuario', 'B'),
        'cargo': ('Cargo', 'C'),
        'modalidad': ('Modalidad', 'D'),
        'calificacion': ('Calificación', 'E'),
        'comunicacion': ('Comunicación', 'F'),
        'dificultad': ('Dificultad', 'G'),
        'tiempo_respuesta': ('Tiempo Respuesta', 'H'),
        'pros': ('Pros', 'I'),
        'contras': ('Contras', 'J'),
        'preguntas': ('Preguntas Entrevista', 'K'),
    }
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Reseñas"
    
    # Estilos
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1E3A8A")
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    row = 1
    
    # ===== ENCABEZADO PRINCIPAL =====
    # Calcular número de columnas para el merge (máximo 11 columnas)
    max_cols = max(len(selected_columns), 8) if selected_columns else 8
    end_header_col = get_column_letter(max_cols)
    ws.merge_cells(f'A{row}:{end_header_col}{row}')
    cell = ws[f'A{row}']
    cell.value = f'REPORTE DE RESEÑAS - {company.name.upper()}'
    cell.font = title_font
    cell.alignment = center_align
    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    row += 1
    
    ws.merge_cells(f'A{row}:{end_header_col}{row}')
    row += 1
    
    ws[f'A{row}'] = 'Período:'
    ws[f'A{row}'].font = subtitle_font
    ws[f'B{row}'] = period_label
    row += 1
    
    ws[f'A{row}'] = 'Fecha de generación:'
    ws[f'A{row}'].font = subtitle_font
    ws[f'B{row}'] = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
    row += 1
    
    ws[f'A{row}'] = 'Generado por:'
    ws[f'A{row}'].font = subtitle_font
    ws[f'B{row}'] = request.user.get_full_name() or request.user.username
    row += 2
    
    # ===== ESTADÍSTICAS GENERALES =====
    # Aplicar filtro de calificación a todas las reseñas si está especificado
    rating_value = None
    if rating_filter:
        try:
            rating_value = int(rating_filter)
            if 1 <= rating_value <= 5:
                # Aplicar filtro a todas las reseñas (aprobadas y rechazadas)
                # Nota: Las reseñas rechazadas también tienen overall_rating
                all_reviews = all_reviews.filter(overall_rating=rating_value)
        except ValueError:
            pass
    
    # Obtener reseñas por estado después de aplicar todos los filtros (período y calificación)
    # Solo consideramos reseñas aprobadas o rechazadas (no pendientes)
    approved_reviews = all_reviews.filter(status='approved')
    rejected_reviews = all_reviews.filter(status='rejected')
    
    # Calcular totales basados en las reseñas filtradas (solo aprobadas y rechazadas)
    approved_count = approved_reviews.count()
    rejected_count = rejected_reviews.count()
    total_reviews = approved_count + rejected_count
    
    # Calcular estadísticas basadas en las reseñas filtradas
    avg_rating = approved_reviews.aggregate(avg=Avg('overall_rating'))['avg'] or 0
    
    # Calcular tasas basadas en el total de reseñas con estado definido (aprobadas + rechazadas)
    total_with_status = approved_count + rejected_count
    
    # Tasa de aprobación: reseñas aprobadas / total de reseñas con estado (aprobadas + rechazadas)
    # Si hay un filtro de calificación, solo considera reseñas de esa calificación
    approval_rate = (approved_count / total_with_status * 100) if total_with_status > 0 else 0
    rejection_rate = (rejected_count / total_with_status * 100) if total_with_status > 0 else 0
    
    # Actualizar etiqueta del período
    if rating_label:
        period_label += rating_label
    
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = 'ESTADÍSTICAS GENERALES'
    cell.font = subtitle_font
    cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.alignment = center_align
    row += 1
    
    stats = [
        ['Métrica', 'Valor'],
        ['Total de reseñas', total_reviews],
        ['Reseñas aprobadas', approved_count],
        ['Reseñas rechazadas', rejected_count],
        ['Calificación promedio (aprobadas)', f'{avg_rating:.2f}/5.0'],
        ['Tasa de aprobación', f'{approval_rate:.1f}%'],
        ['Tasa de rechazo', f'{rejection_rate:.1f}%'],
    ]
    
    for stat_row in stats:
        ws[f'A{row}'] = stat_row[0]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = stat_row[1]
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 1
    
    # ===== DISTRIBUCIÓN POR MODALIDAD =====
    modality_dist = list(approved_reviews.values('modality').annotate(count=Count('id')).order_by('-count'))
    if modality_dist:
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'DISTRIBUCIÓN POR MODALIDAD DE TRABAJO'
        cell.font = subtitle_font
        cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = center_align
        row += 1
        
        ws[f'A{row}'] = 'Modalidad'
        ws[f'B{row}'] = 'Cantidad'
        ws[f'C{row}'] = 'Porcentaje'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1
        
        for item in modality_dist:
            percentage = (item['count'] / approved_reviews.count() * 100) if approved_reviews.count() > 0 else 0
            ws[f'A{row}'] = dict(Review.MODALITY_CHOICES)[item['modality']]
            ws[f'B{row}'] = item['count']
            ws[f'C{row}'] = f'{percentage:.1f}%'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1
        
        row += 1
    
    # ===== DISTRIBUCIÓN POR CALIFICACIÓN =====
    rating_dist = list(approved_reviews.values('overall_rating').annotate(count=Count('id')).order_by('overall_rating'))
    if rating_dist:
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'DISTRIBUCIÓN POR CALIFICACIÓN'
        cell.font = subtitle_font
        cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = center_align
        row += 1
        
        ws[f'A{row}'] = 'Calificación'
        ws[f'B{row}'] = 'Cantidad'
        ws[f'C{row}'] = 'Porcentaje'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1
        
        for item in rating_dist:
            percentage = (item['count'] / approved_reviews.count() * 100) if approved_reviews.count() > 0 else 0
            ws[f'A{row}'] = f"{item['overall_rating']} estrellas"
            ws[f'B{row}'] = item['count']
            ws[f'C{row}'] = f'{percentage:.1f}%'
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1
        
        row += 1
    
    # ===== GRÁFICOS =====
    # Guardar la fila donde empiezan los gráficos (después de las tablas)
    charts_start_row = row
    chart_row_offset = 2  # Espacio antes de los gráficos
    
    # Usar columnas lejanas (M en adelante, columna 13) para los datos de los gráficos
    chart_data_col = 13  # Columna M
    
    # Datos para gráfico de distribución de calificaciones
    if rating_dist:
        rating_data_start_row = charts_start_row + chart_row_offset
        ws.cell(row=rating_data_start_row, column=chart_data_col, value='Calificación')
        ws.cell(row=rating_data_start_row, column=chart_data_col + 1, value='Cantidad')
        rating_data_start_row += 1
        rating_count = len(rating_dist)
        
        for item in rating_dist:
            ws.cell(row=rating_data_start_row, column=chart_data_col, value=f"{item['overall_rating']}⭐")
            ws.cell(row=rating_data_start_row, column=chart_data_col + 1, value=item['count'])
            rating_data_start_row += 1
        
        # Crear gráfico de barras para calificaciones
        if rating_count > 0:
            chart1 = BarChart()
            chart1.type = "col"
            chart1.style = 10
            chart1.title = "Distribución de Calificaciones"
            chart1.y_axis.title = 'Cantidad'
            chart1.x_axis.title = 'Calificación'
            chart1.height = 10
            chart1.width = 15
            
            data = Reference(ws, min_col=chart_data_col + 1, min_row=rating_data_start_row - rating_count, 
                           max_row=rating_data_start_row - 1)
            cats = Reference(ws, min_col=chart_data_col, min_row=rating_data_start_row - rating_count, 
                           max_row=rating_data_start_row - 1)
            chart1.add_data(data, titles_from_data=False)
            chart1.set_categories(cats)
            
            # Colores para las barras (se aplicarán después de agregar el gráfico)
            # Los colores se pueden personalizar pero openpyxl tiene limitaciones
            # Por ahora, dejamos el estilo por defecto que ya es visualmente atractivo
            
            ws.add_chart(chart1, f"A{charts_start_row + chart_row_offset}")
    
    # Datos para gráfico de modalidad
    if modality_dist:
        modality_data_start_row = charts_start_row + chart_row_offset
        ws.cell(row=modality_data_start_row, column=chart_data_col + 3, value='Modalidad')  # Columna P
        ws.cell(row=modality_data_start_row, column=chart_data_col + 4, value='Cantidad')   # Columna Q
        modality_data_start_row += 1
        modality_count = len(modality_dist)
        
        for item in modality_dist:
            modality_name = dict(Review.MODALITY_CHOICES)[item['modality']]
            ws.cell(row=modality_data_start_row, column=chart_data_col + 3, value=modality_name)
            ws.cell(row=modality_data_start_row, column=chart_data_col + 4, value=item['count'])
            modality_data_start_row += 1
        
        # Crear gráfico de pastel para modalidad
        if modality_count > 0:
            chart2 = PieChart()
            chart2.title = "Distribución por Modalidad"
            chart2.height = 10
            chart2.width = 15
            
            data = Reference(ws, min_col=chart_data_col + 4, min_row=modality_data_start_row - modality_count, 
                           max_row=modality_data_start_row - 1)
            cats = Reference(ws, min_col=chart_data_col + 3, min_row=modality_data_start_row - modality_count, 
                           max_row=modality_data_start_row - 1)
            chart2.add_data(data, titles_from_data=False)
            chart2.set_categories(cats)
            
            # Colores para el gráfico de pastel (openpyxl aplica colores automáticamente)
            
            ws.add_chart(chart2, f"J{charts_start_row + chart_row_offset}")
    
    # Datos para gráfico de tiempo de respuesta
    response_time_dist = list(approved_reviews.values('response_time_rating').annotate(count=Count('id')).order_by('response_time_rating'))
    response_time_labels_map = {
        'immediate': 'Inmediata',
        'same_day': 'Mismo día',
        'next_day': 'Al día siguiente',
        'few_days': 'En pocos días',
        'slow': 'Lenta'
    }
    
    if response_time_dist:
        response_data_start_row = charts_start_row + chart_row_offset + 20  # Más abajo
        ws.cell(row=response_data_start_row, column=chart_data_col, value='Tiempo de Respuesta')
        ws.cell(row=response_data_start_row, column=chart_data_col + 1, value='Cantidad')
        response_data_start_row += 1
        response_count = len(response_time_dist)
        
        for item in response_time_dist:
            label = response_time_labels_map.get(item['response_time_rating'], item['response_time_rating'])
            ws.cell(row=response_data_start_row, column=chart_data_col, value=label)
            ws.cell(row=response_data_start_row, column=chart_data_col + 1, value=item['count'])
            response_data_start_row += 1
        
        # Crear gráfico de barras horizontal para tiempo de respuesta
        if response_count > 0:
            chart3 = BarChart()
            chart3.type = "bar"
            chart3.style = 10
            chart3.title = "Distribución de Tiempo de Respuesta"
            chart3.y_axis.title = 'Tiempo de Respuesta'
            chart3.x_axis.title = 'Cantidad'
            chart3.height = 10
            chart3.width = 15
            
            data = Reference(ws, min_col=chart_data_col + 1, min_row=response_data_start_row - response_count, 
                           max_row=response_data_start_row - 1)
            cats = Reference(ws, min_col=chart_data_col, min_row=response_data_start_row - response_count, 
                           max_row=response_data_start_row - 1)
            chart3.add_data(data, titles_from_data=False)
            chart3.set_categories(cats)
            
            # Colores para el gráfico de tiempo de respuesta (openpyxl aplica colores automáticamente)
            
            ws.add_chart(chart3, f"A{charts_start_row + chart_row_offset + 22}")
    
    # Datos para gráfico de comparación mes a mes (solo si hay datos mensuales)
    from django.db.models.functions import TruncMonth
    monthly_trend_data = list(approved_reviews.annotate(
        month=TruncMonth('submission_date')
    ).values('month').annotate(
        count=Count('id'),
        avg_rating=Avg('overall_rating')
    ).order_by('month'))
    
    if len(monthly_trend_data) >= 2:
        monthly_data_start_row = charts_start_row + chart_row_offset + 20
        ws.cell(row=monthly_data_start_row, column=chart_data_col + 3, value='Mes')  # Columna P
        ws.cell(row=monthly_data_start_row, column=chart_data_col + 4, value='Cantidad')  # Columna Q
        ws.cell(row=monthly_data_start_row, column=chart_data_col + 5, value='Calificación')  # Columna R
        monthly_data_start_row += 1
        monthly_count = len(monthly_trend_data)
        
        meses_espanol_short = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
            5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
            9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }
        
        for item in monthly_trend_data:
            month_date = item['month']
            if month_date:
                month_label = f"{meses_espanol_short.get(month_date.month, month_date.strftime('%b'))} {month_date.year}"
                ws.cell(row=monthly_data_start_row, column=chart_data_col + 3, value=month_label)
                ws.cell(row=monthly_data_start_row, column=chart_data_col + 4, value=item['count'])
                ws.cell(row=monthly_data_start_row, column=chart_data_col + 5, value=round(item['avg_rating'], 2) if item['avg_rating'] else 0)
                monthly_data_start_row += 1
        
        # Crear gráfico de línea para comparación mensual
        chart4 = LineChart()
        chart4.title = "Evolución Mensual"
        chart4.style = 13
        chart4.y_axis.title = 'Valor'
        chart4.x_axis.title = 'Mes'
        chart4.height = 10
        chart4.width = 20
        chart4.legend.position = 'b'
        
        # Configurar categorías (meses)
        cats = Reference(ws, min_col=chart_data_col + 3, min_row=monthly_data_start_row - monthly_count, 
                        max_row=monthly_data_start_row - 1)
        chart4.set_categories(cats)
        
        # Datos de cantidad de reseñas (primera serie)
        data1 = Reference(ws, min_col=chart_data_col + 4, min_row=monthly_data_start_row - monthly_count, 
                         max_row=monthly_data_start_row - 1)
        s1 = Series(data1, title="Cantidad de Reseñas")
        chart4.series.append(s1)
        
        # Datos de calificación promedio (segunda serie)
        data2 = Reference(ws, min_col=chart_data_col + 5, min_row=monthly_data_start_row - monthly_count, 
                         max_row=monthly_data_start_row - 1)
        s2 = Series(data2, title="Calificación Promedio")
        chart4.series.append(s2)
        
        ws.add_chart(chart4, f"J{charts_start_row + chart_row_offset + 22}")
    
    # Actualizar row para que las reseñas empiecen después de los gráficos
    row = charts_start_row + 45  # Espacio suficiente para los gráficos
    
    # ===== RESEÑAS APROBADAS =====
    num_cols = len(selected_columns)
    if num_cols > 0:
        start_col = 'A'
        end_col = get_column_letter(num_cols)
        ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
        cell = ws[f'{start_col}{row}']
        cell.value = 'RESEÑAS APROBADAS'
        cell.font = subtitle_font
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = center_align
        row += 1
        
        # Crear headers solo para columnas seleccionadas
        col_idx = 1
        column_positions = {}
        for col_key in selected_columns:
            if col_key in column_map:
                header_name, _ = column_map[col_key]
                column_positions[col_key] = col_idx
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
                col_idx += 1
        row += 1
        
        # Mapeo de datos por columna
        data_map = {
            'fecha': lambda r: r.submission_date.strftime('%d/%m/%Y %H:%M'),
            'usuario': lambda r: r.user_profile.user.get_full_name() or r.user_profile.user.username,
            'cargo': lambda r: r.job_title,
            'modalidad': lambda r: r.get_modality_display(),
            'calificacion': lambda r: f'{r.overall_rating}/5',
            'comunicacion': lambda r: r.get_communication_rating_display(),
            'dificultad': lambda r: r.get_difficulty_rating_display(),
            'tiempo_respuesta': lambda r: r.get_response_time_rating_display(),
            'pros': lambda r: r.pros.replace('\n', ' ').replace('\r', ' ') if r.pros else 'N/A',
            'contras': lambda r: r.cons.replace('\n', ' ').replace('\r', ' ') if r.cons else 'N/A',
            'preguntas': lambda r: r.interview_questions.replace('\n', ' ').replace('\r', ' ') if r.interview_questions else 'N/A',
        }
        
        for review in approved_reviews.order_by('-submission_date'):
            for col_key in selected_columns:
                if col_key in column_positions and col_key in data_map:
                    col_pos = column_positions[col_key]
                    cell = ws.cell(row=row, column=col_pos)
                    cell.value = data_map[col_key](review)
                    cell.border = border
                    if col_key in ['pros', 'contras', 'preguntas']:
                        cell.alignment = wrap_align
            row += 1
    
    row += 1
    
    # ===== RESEÑAS RECHAZADAS =====
    if rejected_reviews.exists():
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = 'RESEÑAS RECHAZADAS'
        cell.font = subtitle_font
        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = center_align
        row += 1
        
        headers = ['Fecha', 'Usuario', 'Cargo', 'Razón de Rechazo', 'Categoría', 'Nivel de Confianza', 'Pros', 'Contras']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        for review in rejected_reviews.order_by('-submission_date'):
            category_display = review.get_verification_category_display() if hasattr(review, 'get_verification_category_display') else review.verification_category
            confidence_display = f'{review.verification_confidence:.2f} ({"Alta" if review.verification_confidence >= 0.8 else "Media" if review.verification_confidence >= 0.6 else "Baja"})' if review.verification_confidence else 'N/A'
            
            ws[f'A{row}'] = review.submission_date.strftime('%d/%m/%Y %H:%M')
            ws[f'B{row}'] = review.user_profile.user.get_full_name() or review.user_profile.user.username
            ws[f'C{row}'] = review.job_title
            ws[f'D{row}'] = (review.verification_reason or 'No especificada').replace('\n', ' ').replace('\r', ' ')
            ws[f'E{row}'] = category_display
            ws[f'F{row}'] = confidence_display
            ws[f'G{row}'] = review.pros.replace('\n', ' ').replace('\r', ' ') if review.pros else 'N/A'
            ws[f'H{row}'] = review.cons.replace('\n', ' ').replace('\r', ' ') if review.cons else 'N/A'
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 7:  # Pros, Contras
                    cell.alignment = wrap_align
            row += 1
    
    # Ajustar anchos de columna según columnas seleccionadas
    width_map = {
        'fecha': 18,
        'usuario': 20,
        'cargo': 25,
        'modalidad': 15,
        'calificacion': 15,
        'comunicacion': 15,
        'dificultad': 15,
        'tiempo_respuesta': 18,
        'pros': 40,
        'contras': 40,
        'preguntas': 40,
    }
    
    col_idx = 1
    for col_key in selected_columns:
        if col_key in column_map and col_key in width_map:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map[col_key]
            col_idx += 1
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar respuesta
    filename = f"reporte_{company.name.replace(' ', '_')}_{period if period != 'all' else 'general'}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response




@login_required
def company_dashboard_view(request):
    """Dashboard para representantes de empresa - Muestra solo estadísticas y gráficos (sin reseñas)"""
    if request.user.profile.role != 'company_rep':
        messages.error(request, 'Acceso no autorizado.')
        return redirect('login')
    
    # Obtener la empresa asociada al usuario
    user_profile = request.user.profile
    if not user_profile.company:
        # Intentar asociar automáticamente la empresa solo si el usuario es "magneto"
        username_lower = request.user.username.lower()
        if username_lower == 'magneto':
            try:
                from companies.models import Company
                company = Company.objects.filter(name__icontains='Magneto Empleos').first()
                if company:
                    user_profile.company = company
                    user_profile.save()
                    messages.success(request, f'✅ Empresa "{company.name}" asociada correctamente.')
                else:
                    messages.error(request, 'No tienes una empresa asociada y no se encontró "Magneto Empleos". Contacta al administrador.')
                    return redirect('my_profile')
            except Exception as e:
                messages.error(request, f'Error al asociar empresa: {str(e)}')
                return redirect('my_profile')
        else:
            messages.error(request, 'No tienes una empresa asociada. Contacta al administrador.')
            return redirect('my_profile')
    
    company = user_profile.company
    
    # Importar modelos necesarios
    from reviews.models import Review
    
    # Obtener todas las reseñas de la empresa (para estadísticas)
    all_company_reviews = Review.objects.filter(company=company)
    
    # Estadísticas para company_rep (siempre calcular, incluso si no hay reseñas)
    total_reviews = all_company_reviews.count()
    approved_count = all_company_reviews.filter(status='approved').count()
    rejected_count = all_company_reviews.filter(status='rejected').count()
    
    approval_rate = (approved_count / total_reviews) * 100 if total_reviews > 0 else 0
    rejection_rate = (rejected_count / total_reviews) * 100 if total_reviews > 0 else 0
    
    # Compromiso de tiempo de respuesta (aprobadas con respuesta rápida)
    compromiso_compliant = all_company_reviews.filter(
        status='approved',
        response_time_rating__in=['immediate', 'same_day']
    ).count()
    compromiso_rate = (compromiso_compliant / approved_count) * 100 if approved_count > 0 else 0
    
    # Calcular promedios (solo si hay reseñas)
    if all_company_reviews.exists():
        avg_overall = all_company_reviews.aggregate(avg=Avg('overall_rating'))['avg'] or 0
    else:
        avg_overall = 0
    
    # Mapear categorías a valores para promediar
    COMM_SCORES = {
        'excellent': 5, 'good': 4, 'regular': 3, 'poor': 2,
    }
    DIFF_SCORES = {
        'very_easy': 1, 'easy': 2, 'moderate': 3, 'difficult': 4, 'very_difficult': 5,
    }
    RESP_SCORES = {
        'immediate': 5, 'same_day': 4, 'next_day': 3, 'few_days': 2, 'slow': 1,
    }
    
    def avg_from_choices(qs, field, scores):
        total = 0
        count = 0
        for val in qs.values_list(field, flat=True).distinct():
            if val in scores:
                val_count = qs.filter(**{field: val}).count()
                total += scores[val] * val_count
                count += val_count
        return (total / count) if count else 0
    
    # Calcular promedios solo si hay reseñas
    if all_company_reviews.exists():
        avg_communication = avg_from_choices(all_company_reviews, 'communication_rating', COMM_SCORES)
        avg_difficulty = avg_from_choices(all_company_reviews, 'difficulty_rating', DIFF_SCORES)
        avg_response_time = avg_from_choices(all_company_reviews, 'response_time_rating', RESP_SCORES)
    else:
        avg_communication = 0
        avg_difficulty = 0
        avg_response_time = 0
    
    # Tendencia mensual
    monthly_trend = all_company_reviews.annotate(
        month=TruncMonth('submission_date')
    ).values('month').annotate(
        count=Count('id'),
        approved=Count('id', filter=Q(status='approved')),
        avg_rating=Avg('overall_rating')
    ).order_by('month')
    
    monthly_ratings = list(monthly_trend)
    
    # Calcular trimestres disponibles basados en los meses con reseñas
    quarters = []
    if monthly_ratings:
        from collections import defaultdict
        quarters_dict = defaultdict(set)
        for item in monthly_ratings:
            month_date = item['month']
            year = month_date.year
            month = month_date.month
            # Determinar trimestre (1-4)
            if month in [1, 2, 3]:
                quarter = 1
                label = f'Q1 {year} (Ene-Mar)'
            elif month in [4, 5, 6]:
                quarter = 2
                label = f'Q2 {year} (Abr-Jun)'
            elif month in [7, 8, 9]:
                quarter = 3
                label = f'Q3 {year} (Jul-Sep)'
            else:  # 10, 11, 12
                quarter = 4
                label = f'Q4 {year} (Oct-Dic)'
            quarters_dict[(year, quarter)] = label
        
        # Convertir a lista ordenada
        for (year, quarter), label in sorted(quarters_dict.items()):
            quarters.append({
                'year': year,
                'quarter': quarter,
                'label': label
            })
    
    # Datos para gráficos
    chart_data = {
        'ratings': {},
        'modality': {},
        'response_time_distribution': {},
        'monthly_comparison': {},
        'strengths_weaknesses': {}
    }
    
    if all_company_reviews.exists():
        # Distribución de calificaciones
        for i in range(1, 6):
            count = all_company_reviews.filter(overall_rating=i).count()
            chart_data['ratings'][f'{i} estrella{"s" if i > 1 else ""}'] = count
        
        # Distribución por modalidad
        for modality, _ in Review.MODALITY_CHOICES:
            count = all_company_reviews.filter(modality=modality).count()
            if count > 0:
                chart_data['modality'][modality.title()] = count
        
        # Distribución de tiempo de respuesta
        response_time_labels = {
            'immediate': 'Inmediata',
            'same_day': 'Mismo día',
            'next_day': 'Al día siguiente',
            'few_days': 'En pocos días',
            'slow': 'Lenta'
        }
        for response_time, _ in Review.RESPONSE_TIME_CHOICES:
            count = all_company_reviews.filter(response_time_rating=response_time).count()
            if count > 0:
                chart_data['response_time_distribution'][response_time_labels[response_time]] = count
        
        # Comparación mensual
        if monthly_ratings:
            chart_data['monthly_comparison'] = {
                str(item['month']): {
                    'count': item['count'],
                    'approved': item['approved'],
                    'avg_rating': float(item['avg_rating']) if item['avg_rating'] else 0
                }
                for item in monthly_ratings
            }
        
        # Gráfico de Fortalezas y Debilidades
        strengths_weaknesses = {}
        aspects = {
            'communication_rating': 'Comunicación',
            'difficulty_rating': 'Dificultad del Proceso', 
            'response_time_rating': 'Tiempo de Respuesta',
            'overall_rating': 'Calificación General'
        }
        
        for field, label in aspects.items():
            if field == 'overall_rating':
                avg_rating = all_company_reviews.aggregate(avg=Avg(field))['avg']
                if avg_rating is not None:
                    strengths_weaknesses[label] = round(avg_rating, 1)
            else:
                rating_mapping = {
                    'communication_rating': {
                        'excellent': 5, 'good': 4, 'regular': 3, 'poor': 2
                    },
                    'difficulty_rating': {
                        'very_easy': 1, 'easy': 2, 'moderate': 3, 'difficult': 4, 'very_difficult': 5
                    },
                    'response_time_rating': {
                        'immediate': 5, 'same_day': 4, 'next_day': 3, 'few_days': 2, 'slow': 1
                    }
                }
                
                total_score = 0
                count = 0
                for review in all_company_reviews:
                    rating_value = rating_mapping[field].get(getattr(review, field), 0)
                    if rating_value > 0:
                        total_score += rating_value
                        count += 1
                
                if count > 0:
                    avg_rating = total_score / count
                    strengths_weaknesses[label] = round(avg_rating, 1)
        
        chart_data['strengths_weaknesses'] = strengths_weaknesses
    
    # Generar recomendaciones inteligentes (solo si hay reseñas)
    # Siempre mostrar máximo 2 recomendaciones
    recommendations = []
    
    if total_reviews > 0:
        # Priorizar recomendaciones según importancia
        if avg_overall < 3.5:
            recommendations.append({
                'type': 'danger',
                'icon': 'fas fa-star',
                'title': 'Calificación General Baja',
                'description': f'Tu calificación promedio es {avg_overall:.1f}/5. Necesitas mejorar varios aspectos.',
            })
        
        if rejection_rate > 30:
            recommendations.append({
                'type': 'danger',
                'icon': 'fas fa-exclamation-triangle',
                'title': 'Alta Tasa de Rechazo',
                'description': f'El {rejection_rate:.1f}% de tus reseñas son rechazadas. Esto puede indicar problemas en tu proceso.',
            })
        
        if avg_communication < 4 and len(recommendations) < 2:
            recommendations.append({
                'type': 'warning',
                'icon': 'fas fa-comments',
                'title': 'Mejorar Comunicación',
                'description': f'Tu calificación de comunicación es {avg_communication:.1f}/5. Considera ser más claro y transparente en tus respuestas a los candidatos.',
            })
        
        if avg_response_time < 4 and len(recommendations) < 2:
            recommendations.append({
                'type': 'info',
                'icon': 'fas fa-clock',
                'title': 'Acelerar Respuestas',
                'description': f'Tu tiempo de respuesta promedio es {avg_response_time:.1f}/5. Los candidatos valoran respuestas rápidas.',
            })
        
        if approval_rate < 70 and len(recommendations) < 2:
            recommendations.append({
                'type': 'warning',
                'icon': 'fas fa-chart-line',
                'title': 'Baja Tasa de Aprobación',
                'description': f'Solo el {approval_rate:.1f}% de tus reseñas son aprobadas. Mejora la calidad de tus procesos.',
            })
        
        if compromiso_rate < 50 and len(recommendations) < 2:
            recommendations.append({
                'type': 'info',
                'icon': 'fas fa-bolt',
                'title': 'Mejorar Velocidad de Respuesta',
                'description': f'Solo el {compromiso_rate:.1f}% de tus reseñas aprobadas tienen respuesta rápida.',
            })
        
        # Si no hay recomendaciones críticas, mostrar positivas
        if not recommendations:
            recommendations.append({
                'type': 'success',
                'icon': 'fas fa-check-circle',
                'title': '¡Excelente Trabajo!',
                'description': 'Tus métricas están en buen nivel. Sigue así y continúa mejorando.',
            })
            recommendations.append({
                'type': 'info',
                'icon': 'fas fa-chart-line',
                'title': 'Mantén el Buen Rendimiento',
                'description': 'Continúa monitoreando tus métricas y busca oportunidades de mejora continua.',
            })
        # Si solo hay 1 recomendación, agregar una segunda positiva o informativa
        elif len(recommendations) == 1:
            if avg_overall >= 4.0:
                recommendations.append({
                    'type': 'success',
                    'icon': 'fas fa-check-circle',
                    'title': '¡Buen Rendimiento!',
                    'description': 'Tus calificaciones están por encima del promedio. Continúa mejorando.',
                })
            elif avg_communication >= 4.0:
                recommendations.append({
                    'type': 'info',
                    'icon': 'fas fa-comments',
                    'title': 'Buena Comunicación',
                    'description': 'Tu comunicación es valorada positivamente por los candidatos.',
                })
            elif avg_response_time >= 4.0:
                recommendations.append({
                    'type': 'info',
                    'icon': 'fas fa-clock',
                    'title': 'Respuestas Rápidas',
                    'description': 'Estás respondiendo rápidamente a los candidatos. ¡Sigue así!',
                })
            else:
                recommendations.append({
                    'type': 'info',
                    'icon': 'fas fa-chart-line',
                    'title': 'Continúa Mejorando',
                    'description': 'Monitorea tus métricas regularmente y busca oportunidades de mejora continua.',
                })
    else:
        # Si no hay reseñas, mostrar 2 recomendaciones iniciales
        recommendations.append({
            'type': 'info',
            'icon': 'fas fa-info-circle',
            'title': 'Comienza a recibir reseñas',
            'description': 'Aún no hay reseñas para tu empresa. Cuando los candidatos compartan sus experiencias, verás estadísticas aquí.',
        })
        recommendations.append({
            'type': 'success',
            'icon': 'fas fa-rocket',
            'title': 'Prepara tu proceso',
            'description': 'Asegúrate de tener un proceso de selección claro y transparente para recibir buenas calificaciones.',
        })
    
    # Limitar a exactamente 2 recomendaciones
    recommendations = recommendations[:2]
    
    # Datos para comparación mes a mes
    monthly_comparison = []
    if len(monthly_ratings) >= 2:
        for i in range(1, len(monthly_ratings)):
            prev_month = monthly_ratings[i-1]
            curr_month = monthly_ratings[i]
            
            # Calcular rechazadas para cada mes
            prev_rejected = prev_month['count'] - prev_month['approved']
            curr_rejected = curr_month['count'] - curr_month['approved']
            
            rating_change = curr_month['avg_rating'] - prev_month['avg_rating']
            count_change = curr_month['count'] - prev_month['count']
            approval_change = curr_month['approved'] - prev_month['approved']
            rejection_change = curr_rejected - prev_rejected
            
            monthly_comparison.append({
                'month': curr_month['month'],
                'prev_month': prev_month['month'],
                'prev_rating': prev_month['avg_rating'],
                'curr_rating': curr_month['avg_rating'],
                'rating_change': rating_change,
                'prev_count': prev_month['count'],
                'curr_count': curr_month['count'],
                'count_change': count_change,
                'prev_approved': prev_month['approved'],
                'curr_approved': curr_month['approved'],
                'approval_change': approval_change,
                'prev_rejected': prev_rejected,
                'curr_rejected': curr_rejected,
                'rejection_change': rejection_change,
            })
    
    role_kpis = {
        'role': 'company_rep',
        'avg_overall': avg_overall,
        'total_reviews': total_reviews,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'approval_rate': round(approval_rate, 1),
        'rejection_rate': round(rejection_rate, 1),
        'compromiso_rate': round(compromiso_rate, 1),
        'monthly_trend': monthly_ratings,
        'monthly_comparison': monthly_comparison,
        'avg_communication': avg_communication,
        'avg_difficulty': avg_difficulty,
        'avg_response_time': avg_response_time,
        'recommendations': recommendations,
        'quarters': quarters,
    }
    
    context = {
        'company': company,
        'chart_data': chart_data,
        'role_kpis': role_kpis,
    }
    
    return render(request, 'companies/company_dashboard.html', context)


@login_required
def company_reviews_view(request):
    """Vista para que company_rep vea las reseñas de su empresa"""
    if request.user.profile.role != 'company_rep':
        messages.error(request, 'Acceso no autorizado.')
        return redirect('login')
    
    # Obtener la empresa asociada al usuario
    user_profile = request.user.profile
    if not user_profile.company:
        messages.error(request, 'No tienes una empresa asociada. Contacta al administrador.')
        return redirect('login')
    
    company = user_profile.company
    
    # Importar modelos necesarios
    from reviews.models import Review
    
    # Obtener todas las reseñas aprobadas SIN filtrar (para contar el total)
    all_approved_reviews = Review.objects.filter(
        company=company,
        status='approved'
    ).select_related('user_profile__user')
    
    # Obtener reseñas con filtros
    approved_reviews = all_approved_reviews
    
    # Aplicar filtros desde parámetros GET
    rating_filter = request.GET.get('rating')
    modality_filter = request.GET.get('modality')
    sort_by = request.GET.get('sort', 'recent')
    
    if rating_filter:
        try:
            rating_value = int(rating_filter)
            if 1 <= rating_value <= 5:
                approved_reviews = approved_reviews.filter(overall_rating=rating_value)
        except ValueError:
            pass
    
    if modality_filter:
        approved_reviews = approved_reviews.filter(modality=modality_filter)
    
    # Aplicar ordenamiento
    if sort_by == 'recent':
        approved_reviews = approved_reviews.order_by('-submission_date')
    elif sort_by == 'oldest':
        approved_reviews = approved_reviews.order_by('submission_date')
    elif sort_by == 'highest':
        approved_reviews = approved_reviews.order_by('-overall_rating', '-submission_date')
    elif sort_by == 'lowest':
        approved_reviews = approved_reviews.order_by('overall_rating', '-submission_date')
    else:
        approved_reviews = approved_reviews.order_by('-submission_date')
    
    rejected_reviews = Review.objects.filter(
        company=company,
        status='rejected'
    ).select_related('user_profile__user')
    
    # Contar el total de reseñas aprobadas SIN filtrar
    total_approved_reviews_count = all_approved_reviews.count()
    rejected_reviews_count = rejected_reviews.count()
    
    # Verificar si hay filtros activos
    has_active_filters = bool(rating_filter or modality_filter)
    
    # Pasar filtros actuales al contexto
    current_filters = {
        'rating': rating_filter,
        'modality': modality_filter,
        'sort': sort_by,
    }
    
    context = {
        'company': company,
        'approved_reviews': approved_reviews,
        'rejected_reviews': rejected_reviews,
        'rejected_reviews_count': rejected_reviews_count,
        'total_approved_reviews_count': total_approved_reviews_count,
        'has_active_filters': has_active_filters,
        'current_filters': current_filters,
    }
    
    return render(request, 'companies/company_reviews.html', context)


@login_required
def staff_dashboard_view(request):
    """Dashboard para staff - Muestra todas las empresas con opciones de administración"""
    if not (request.user.is_staff or request.user.profile.role == 'staff'):
        messages.error(request, 'Acceso no autorizado. Solo el staff puede acceder.')
        return redirect('dashboard')
    
    # Importar modelos necesarios
    from reviews.models import Review
    
    # Búsqueda
    search_query = request.GET.get('search', '')
    companies = Company.objects.all().prefetch_related('reviews').order_by('name')
    
    if search_query:
        companies = companies.filter(
            Q(name__icontains=search_query) |
            Q(sector__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    # Estadísticas
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(is_active=True).count()
    inactive_companies = Company.objects.filter(is_active=False).count()
    total_reviews = Review.objects.count()
    total_candidates = UserProfile.objects.filter(role='candidate').count()
    
    # Agregar info a cada empresa
    for company in companies:
        company_reviews = company.reviews.all()
        company.total_reviews = company_reviews.count()
        
        company_rated_reviews = company_reviews.filter(overall_rating__isnull=False)
        if company_rated_reviews.exists():
            company.avg_rating = company_reviews.aggregate(avg=Avg('overall_rating'))['avg']
        else:
            company.avg_rating = 0
    
    context = {
        'companies': companies,
        'total_companies': total_companies,
        'active_companies': active_companies,
        'inactive_companies': inactive_companies,
        'total_reviews': total_reviews,
        'total_candidates': total_candidates,
        'search_query': search_query,
    }
    
    return render(request, 'companies/staff_dashboard.html', context)


@login_required
def edit_company_view(request, company_id):
    """Vista para editar información de empresa (para company_rep y staff)"""
    company = get_object_or_404(Company, id=company_id)
    
    # Verificar permisos
    is_staff = request.user.is_staff or request.user.profile.role == 'staff'
    is_company_rep = request.user.profile.role == 'company_rep' and request.user.profile.company and request.user.profile.company.id == company.id
    
    if not (is_staff or is_company_rep):
        messages.error(request, 'No tienes permisos para editar empresas.')
        return redirect('company_detail', company_id=company.id)
    
    if request.method == 'POST':
        form = CompanyEditForm(request.POST, request.FILES, instance=company, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Información de la empresa actualizada exitosamente.')
            if is_staff:
                return redirect('staff_dashboard')
            elif request.user.profile.role == 'company_rep':
                return redirect('company_dashboard')
            return redirect('company_detail', company_id=company.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = CompanyEditForm(instance=company, user=request.user)
    
    context = {
        'form': form,
        'company': company,
        'is_staff': is_staff,
    }
    
    return render(request, 'companies/edit_company.html', context)


@login_required
def toggle_company_status_view(request, company_id):
    """Activar/desactivar empresa (solo staff)"""
    if not (request.user.is_staff or request.user.profile.role == 'staff'):
        messages.error(request, 'Acceso no autorizado.')
        return redirect('dashboard')
    
    company = get_object_or_404(Company, id=company_id)
    company.is_active = not company.is_active
    company.save()
    
    status = 'activada' if company.is_active else 'desactivada'
    messages.success(request, f'Empresa {company.name} {status} exitosamente.')
    
    return redirect('staff_dashboard')


@login_required
def delete_company_view(request, company_id):
    """Borrar empresa (solo staff)"""
    if not (request.user.is_staff or request.user.profile.role == 'staff'):
        messages.error(request, 'Acceso no autorizado.')
        return redirect('dashboard')
    
    company = get_object_or_404(Company, id=company_id)
    company_name = company.name
    company.delete()
    
    messages.success(request, f'Empresa {company_name} borrada exitosamente.')
    return redirect('staff_dashboard')


@login_required
def create_company_view(request):
    """Crear nueva empresa (solo staff)"""
    if not (request.user.is_staff or request.user.profile.role == 'staff'):
        messages.error(request, 'Acceso no autorizado.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CompanyEditForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            company = form.save()
            messages.success(request, f'Empresa {company.name} creada exitosamente.')
            return redirect('staff_dashboard')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = CompanyEditForm(user=request.user)
    
    context = {
        'form': form,
        'is_new': True,
    }
    
    return render(request, 'companies/edit_company.html', context)


@login_required
def delete_review_view(request, review_id):
    """Borrar reseña (solo staff)"""
    if not (request.user.is_staff or request.user.profile.role == 'staff'):
        messages.error(request, 'Acceso no autorizado.')
        return redirect('dashboard')
    
    from reviews.models import Review
    review = get_object_or_404(Review, id=review_id)
    company_id = review.company.id
    review.delete()
    
    messages.success(request, 'Reseña borrada exitosamente.')
    return redirect('company_detail', company_id=company_id)